"""
FINEP Loader — Proj__Crédito Descentralizado

Lê a aba Proj__Crédito_Descentralizado do arquivo Contratacao.xlsx mais recente
(ou do manifest_id informado) e popula finep.projetos_credito_descentralizado.

Estrutura da aba:
  - 3.035 linhas, 10 colunas
  - Header: linha 6 | Dados: linha 7..
  - contrato = convênio-quadro FINEP↔banco (até 709 sub-operações por contrato)

Estratégia de carga:
  - DELETE + INSERT por manifest_id: idempotente.
"""

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)

SHEET_NAME = "Proj__Crédito_Descentralizado"
HEADER_ROW = 6
DATA_START  = 7

# Mapeamento col Excel (0-based) → campo da tabela
COL_MAP = {
    0:  "data_assinatura",
    1:  "contrato",
    2:  "proponente",
    3:  "cnpj_proponente",
    4:  "uf",
    5:  "valor_financiado",
    6:  "valor_liberado",
    7:  "contrapartida",
    8:  "outros_recursos",
    9:  "agente_financeiro",
}

INSERT_SQL = """
    INSERT INTO finep.projetos_credito_descentralizado (
        data_assinatura, contrato, proponente, cnpj_proponente,
        uf, valor_financiado, valor_liberado,
        contrapartida, outros_recursos,
        agente_financeiro, manifest_id
    ) VALUES (
        %(data_assinatura)s, %(contrato)s, %(proponente)s, %(cnpj_proponente)s,
        %(uf)s, %(valor_financiado)s, %(valor_liberado)s,
        %(contrapartida)s, %(outros_recursos)s,
        %(agente_financeiro)s, %(manifest_id)s
    )
"""


# ---------------------------------------------------------------------------
# Banco de dados
# ---------------------------------------------------------------------------

def _get_db_connection():
    return psycopg2.connect(
        host=os.getenv("POSTGRES_HOST", "postgres"),
        port=int(os.getenv("POSTGRES_PORT", "5432")),
        database=os.getenv("POSTGRES_DB", "osint_metadata"),
        user=os.getenv("POSTGRES_USER", "osint_admin"),
        password=os.getenv("POSTGRES_PASSWORD", "osint_secure_password"),
    )


def get_latest_manifest(dataset_type: str = "contratacao") -> Optional[dict]:
    """Retorna manifest_id e local_path do download mais recente validado."""
    conn = _get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("""
                SELECT id, local_path, download_date
                FROM finep.download_manifest
                WHERE dataset_type = %s
                  AND processing_status IN ('downloaded', 'validated', 'loaded')
                  AND local_path IS NOT NULL
                ORDER BY download_date DESC
                LIMIT 1
            """, (dataset_type,))
            row = cur.fetchone()
            return dict(row) if row else None
    finally:
        conn.close()


def mark_manifest_loaded(manifest_id: int, rows_loaded: int) -> None:
    """Atualiza o manifesto indicando que esta aba foi carregada."""
    conn = _get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE finep.download_manifest
                SET processing_status = 'loaded',
                    rows_count        = COALESCE(rows_count, 0) + %s,
                    updated_at        = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (rows_loaded, manifest_id))
            conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_numeric(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _clean_date(val):
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date()
    return None


# ---------------------------------------------------------------------------
# Leitura do xlsx
# ---------------------------------------------------------------------------

def iter_rows(xlsx_path: Path):
    """
    Lê a aba Proj__Crédito_Descentralizado e gera dicts prontos para INSERT.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            if all(v is None for v in row):
                continue

            record = {
                "data_assinatura":   _clean_date(row[0]),
                "contrato":          _clean_str(row[1]),
                "proponente":        _clean_str(row[2]),
                "cnpj_proponente":   _clean_str(row[3]),
                "uf":                _clean_str(row[4]),
                "valor_financiado":  _clean_numeric(row[5]),
                "valor_liberado":    _clean_numeric(row[6]),
                "contrapartida":     _clean_numeric(row[7]),
                "outros_recursos":   _clean_numeric(row[8]),
                "agente_financeiro": _clean_str(row[9]) if len(row) > 9 else None,
                "manifest_id":       None,
            }

            # Ignora linhas de separação/subtotal sem contrato
            if not record["contrato"]:
                continue

            yield record
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Carga principal
# ---------------------------------------------------------------------------

def load_credito_descentralizado(
    manifest_id: Optional[int] = None,
    xlsx_path: Optional[str] = None,
    batch_size: int = 500,
) -> dict:
    """
    Carrega a aba Proj__Crédito_Descentralizado para o PostgreSQL.

    Args:
        manifest_id: ID do registro em finep.download_manifest.
                     Se None, usa o download mais recente.
        xlsx_path:   Caminho do arquivo. Se None, busca no manifest.
        batch_size:  Linhas por batch de INSERT.

    Returns:
        dict com manifest_id, rows_loaded, duration_seconds
    """
    t_start = datetime.now(tz=timezone.utc)

    if manifest_id is None or xlsx_path is None:
        manifest = get_latest_manifest("contratacao")
        if not manifest:
            raise RuntimeError("Nenhum download de contratacao encontrado no manifesto.")
        manifest_id = manifest_id or manifest["id"]
        xlsx_path   = xlsx_path   or manifest["local_path"]

    logger.info(f"Carregando Crédito Descentralizado | manifest_id={manifest_id} | {xlsx_path}")

    conn = _get_db_connection()
    rows_loaded = 0

    try:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM finep.projetos_credito_descentralizado WHERE manifest_id = %s",
                (manifest_id,),
            )
            deleted = cur.rowcount
            if deleted:
                logger.info(f"Removidos {deleted} registros anteriores do manifest_id={manifest_id}")

            batch = []
            for record in iter_rows(Path(xlsx_path)):
                record["manifest_id"] = manifest_id
                batch.append(record)

                if len(batch) >= batch_size:
                    psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                    rows_loaded += len(batch)
                    batch = []
                    logger.debug(f"  {rows_loaded} linhas inseridas...")

            if batch:
                psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                rows_loaded += len(batch)

        conn.commit()
        logger.info(f"✓ {rows_loaded} linhas carregadas com sucesso")

    except Exception as exc:
        conn.rollback()
        logger.error(f"Falha na carga: {exc}")
        raise
    finally:
        conn.close()

    duration = (datetime.now(tz=timezone.utc) - t_start).total_seconds()
    mark_manifest_loaded(manifest_id, rows_loaded)

    return {
        "manifest_id":      manifest_id,
        "rows_loaded":      rows_loaded,
        "duration_seconds": round(duration, 2),
        "xlsx_path":        xlsx_path,
    }
