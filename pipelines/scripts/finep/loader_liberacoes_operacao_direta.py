"""
FINEP Loader — Liberações Operação Direta

Lê a aba Projetos_Operação_Direta do Liberacao.xlsx e popula
finep.liberacoes_operacao_direta.

Modelo normalizado: somente os campos delta de cada liberação.
Contexto do contrato via JOIN com finep.projetos_operacao_direta
usando (contrato, ref).

Mapeamento de colunas (0-based):
  2  → ref
  3  → contrato
  14 → num_parcela     (Nº Parcela)
  15 → num_liberacao   (Nº Liberação)
  16 → data_liberacao  (Data Liberação)
  17 → valor_liberado  (Valor Liberado)
"""

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)

SHEET_NAME = "Projetos_Operação_Direta"
HEADER_ROW = 6
DATA_START  = 7

INSERT_SQL = """
    INSERT INTO finep.liberacoes_operacao_direta (
        contrato, ref,
        num_parcela, num_liberacao,
        data_liberacao, valor_liberado,
        manifest_id
    ) VALUES (
        %(contrato)s, %(ref)s,
        %(num_parcela)s, %(num_liberacao)s,
        %(data_liberacao)s, %(valor_liberado)s,
        %(manifest_id)s
    )
"""


def _get_db_connection():
    return psycopg2.connect(
        host=os.getenv("POSTGRES_HOST", "postgres"),
        port=int(os.getenv("POSTGRES_PORT", "5432")),
        database=os.getenv("POSTGRES_DB", "osint_metadata"),
        user=os.getenv("POSTGRES_USER", "osint_admin"),
        password=os.getenv("POSTGRES_PASSWORD", "osint_secure_password"),
    )


def get_latest_manifest() -> Optional[dict]:
    conn = _get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("""
                SELECT id, local_path, download_date
                FROM finep.download_manifest
                WHERE dataset_type = 'liberacao'
                  AND processing_status IN ('downloaded', 'validated', 'loaded')
                  AND local_path IS NOT NULL
                ORDER BY download_date DESC
                LIMIT 1
            """)
            row = cur.fetchone()
            return dict(row) if row else None
    finally:
        conn.close()


def mark_manifest_loaded(manifest_id: int, rows_loaded: int) -> None:
    conn = _get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE finep.download_manifest
                SET processing_status = 'loaded',
                    rows_count        = COALESCE(rows_count, 0) + %s,
                    updated_at        = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (rows_loaded, manifest_id))
            conn.commit()
    finally:
        conn.close()


def _clean_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_numeric(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _clean_smallint(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _clean_date(val):
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date()
    return None


def iter_rows(xlsx_path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            if all(v is None for v in row):
                continue

            contrato = _clean_str(row[3])
            if not contrato:
                continue

            yield {
                "ref":            _clean_str(row[2]),
                "contrato":       contrato,
                "num_parcela":    _clean_smallint(row[14]),
                "num_liberacao":  _clean_smallint(row[15]),
                "data_liberacao": _clean_date(row[16]),
                "valor_liberado": _clean_numeric(row[17]),
                "manifest_id":    None,
            }
    finally:
        wb.close()


def load_liberacoes_operacao_direta(
    manifest_id: Optional[int] = None,
    xlsx_path: Optional[str] = None,
    batch_size: int = 1000,
) -> dict:
    t_start = datetime.now(tz=timezone.utc)

    if manifest_id is None or xlsx_path is None:
        manifest = get_latest_manifest()
        if not manifest:
            raise RuntimeError("Nenhum download de liberacao encontrado no manifesto.")
        manifest_id = manifest_id or manifest["id"]
        xlsx_path   = xlsx_path   or manifest["local_path"]

    logger.info(f"Carregando liberacoes_operacao_direta | manifest_id={manifest_id} | {xlsx_path}")

    conn = _get_db_connection()
    rows_loaded = 0

    try:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM finep.liberacoes_operacao_direta WHERE manifest_id = %s",
                (manifest_id,),
            )
            deleted = cur.rowcount
            if deleted:
                logger.info(f"Removidos {deleted} registros anteriores do manifest_id={manifest_id}")

            batch = []
            for record in iter_rows(Path(xlsx_path)):
                record["manifest_id"] = manifest_id
                batch.append(record)
                if len(batch) >= batch_size:
                    psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                    rows_loaded += len(batch)
                    batch = []
                    logger.debug(f"  {rows_loaded} linhas inseridas...")

            if batch:
                psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                rows_loaded += len(batch)

        conn.commit()
        logger.info(f"✓ {rows_loaded} linhas carregadas com sucesso")

    except Exception as exc:
        conn.rollback()
        logger.error(f"Falha na carga: {exc}")
        raise
    finally:
        conn.close()

    duration = (datetime.now(tz=timezone.utc) - t_start).total_seconds()
    mark_manifest_loaded(manifest_id, rows_loaded)

    return {
        "manifest_id":      manifest_id,
        "rows_loaded":      rows_loaded,
        "duration_seconds": round(duration, 2),
        "xlsx_path":        xlsx_path,
    }
