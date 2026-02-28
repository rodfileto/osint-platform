"""
FINEP Loader — Projetos Operação Direta

Lê a aba Projetos_Operação_Direta do arquivo Contratacao.xlsx mais recente
(ou do manifest_id informado) e popula finep.projetos_operacao_direta.

Estratégia de carga:
  - DELETE + INSERT por manifest_id: garante idempotência.
    Re-executar a task apaga os dados da carga anterior daquele arquivo
    e insere novamente — seguro para re-runs.
  - Nunca apaga dados de outros downloads (manifest_ids diferentes).
"""

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)

SHEET_NAME = "Projetos_Operação_Direta"
HEADER_ROW = 6    # linha do cabeçalho no xlsx (1-based)
DATA_START  = 7   # primeira linha de dados

# Mapeamento col Excel (0-based) → campo da tabela
COL_MAP = {
    0:  "instrumento",
    1:  "demanda",
    2:  "ref",
    3:  "contrato",
    4:  "data_assinatura",
    5:  "prazo_execucao_original",
    6:  "prazo_execucao",
    7:  "titulo",
    8:  "proponente",
    9:  "cnpj_proponente",
    10: "executor",
    11: "cnpj_executor",
    12: "municipio",
    13: "uf",
    14: "valor_finep",
    15: "status",
    16: "contrapartida_financeira",
    17: "contrapartida_nao_financeira",
    18: "valor_pago",
    19: "intervenientes",
    20: "aporte_financeiro_interv",
    21: "aporte_nao_financeiro_interv",
    22: "resumo_publicavel",
}

INSERT_SQL = """
    INSERT INTO finep.projetos_operacao_direta (
        instrumento, demanda, ref, contrato,
        data_assinatura, prazo_execucao_original, prazo_execucao,
        titulo, status,
        proponente, cnpj_proponente,
        executor, cnpj_executor,
        municipio, uf,
        valor_finep, contrapartida_financeira, contrapartida_nao_financeira,
        valor_pago, intervenientes,
        aporte_financeiro_interv, aporte_nao_financeiro_interv,
        resumo_publicavel, manifest_id
    ) VALUES (
        %(instrumento)s, %(demanda)s, %(ref)s, %(contrato)s,
        %(data_assinatura)s, %(prazo_execucao_original)s, %(prazo_execucao)s,
        %(titulo)s, %(status)s,
        %(proponente)s, %(cnpj_proponente)s,
        %(executor)s, %(cnpj_executor)s,
        %(municipio)s, %(uf)s,
        %(valor_finep)s, %(contrapartida_financeira)s, %(contrapartida_nao_financeira)s,
        %(valor_pago)s, %(intervenientes)s,
        %(aporte_financeiro_interv)s, %(aporte_nao_financeiro_interv)s,
        %(resumo_publicavel)s, %(manifest_id)s
    )
"""


# ---------------------------------------------------------------------------
# Banco de dados
# ---------------------------------------------------------------------------

def _get_db_connection():
    return psycopg2.connect(
        host=os.getenv("POSTGRES_HOST", "postgres"),
        port=int(os.getenv("POSTGRES_PORT", "5432")),
        database=os.getenv("POSTGRES_DB", "osint_metadata"),
        user=os.getenv("POSTGRES_USER", "osint_admin"),
        password=os.getenv("POSTGRES_PASSWORD", "osint_secure_password"),
    )


def get_latest_manifest(dataset_type: str = "contratacao") -> Optional[dict]:
    """Retorna o manifest_id e local_path do download mais recente validado."""
    conn = _get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("""
                SELECT id, local_path, download_date
                FROM finep.download_manifest
                WHERE dataset_type = %s
                  AND processing_status IN ('downloaded', 'validated')
                  AND local_path IS NOT NULL
                ORDER BY download_date DESC
                LIMIT 1
            """, (dataset_type,))
            row = cur.fetchone()
            return dict(row) if row else None
    finally:
        conn.close()


def mark_manifest_loaded(manifest_id: int, rows_loaded: int) -> None:
    """Atualiza o manifesto indicando que os dados foram carregados."""
    conn = _get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE finep.download_manifest
                SET processing_status = 'loaded',
                    rows_count        = %s,
                    updated_at        = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (rows_loaded, manifest_id))
            conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Leitura do xlsx
# ---------------------------------------------------------------------------

def _clean_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_numeric(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _clean_int(val) -> Optional[int]:
    if val is None:
        return 0
    try:
        return int(val)
    except (TypeError, ValueError):
        return 0


def _clean_date(val):
    """Aceita datetime do openpyxl ou None."""
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date()
    return None


def iter_rows(xlsx_path: Path):
    """
    Lê a aba Projetos_Operação_Direta e gera dicts prontos para INSERT.
    Ignora linhas completamente vazias.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            # Ignorar linhas em branco
            if all(v is None for v in row):
                continue

            record = {
                "instrumento":                _clean_str(row[0]),
                "demanda":                    _clean_str(row[1]),
                "ref":                        _clean_str(row[2]),
                "contrato":                   _clean_str(row[3]),
                "data_assinatura":            _clean_date(row[4]),
                "prazo_execucao_original":    _clean_date(row[5]),
                "prazo_execucao":             _clean_date(row[6]),
                "titulo":                     _clean_str(row[7]),
                "proponente":                 _clean_str(row[8]),
                "cnpj_proponente":            _clean_str(row[9]),
                "executor":                   _clean_str(row[10]),
                "cnpj_executor":              _clean_str(row[11]),
                "municipio":                  _clean_str(row[12]),
                "uf":                         _clean_str(row[13]),
                "valor_finep":                _clean_numeric(row[14]),
                "status":                     _clean_str(row[15]),
                "contrapartida_financeira":   _clean_numeric(row[16]),
                "contrapartida_nao_financeira": _clean_numeric(row[17]),
                "valor_pago":                 _clean_numeric(row[18]),
                "intervenientes":             _clean_int(row[19]),
                "aporte_financeiro_interv":   _clean_numeric(row[20]),
                "aporte_nao_financeiro_interv": _clean_numeric(row[21]),
                "resumo_publicavel":          _clean_str(row[22]) if len(row) > 22 else None,
                "manifest_id":                None,   # preenchido antes do INSERT
            }

            # Descartar linhas sem instrumento (linhas de separação/subtotal)
            if not record["instrumento"]:
                continue

            yield record
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Carga principal
# ---------------------------------------------------------------------------

def load_operacao_direta(
    manifest_id: Optional[int] = None,
    xlsx_path: Optional[str] = None,
    batch_size: int = 500,
) -> dict:
    """
    Carrega a aba Projetos_Operação_Direta para o PostgreSQL.

    Args:
        manifest_id: ID do registro em finep.download_manifest.
                     Se None, usa o download mais recente.
        xlsx_path:   Caminho do arquivo. Se None, busca no manifest.
        batch_size:  Linhas por batch de INSERT.

    Returns:
        dict com manifest_id, rows_loaded, duration_seconds
    """
    t_start = datetime.now(tz=timezone.utc)

    # Resolver manifest e caminho do arquivo
    if manifest_id is None or xlsx_path is None:
        manifest = get_latest_manifest("contratacao")
        if not manifest:
            raise RuntimeError("Nenhum download de contratacao encontrado no manifesto.")
        manifest_id = manifest_id or manifest["id"]
        xlsx_path = xlsx_path or manifest["local_path"]

    logger.info(f"Carregando Projetos_Operação_Direta | manifest_id={manifest_id} | {xlsx_path}")

    conn = _get_db_connection()
    rows_loaded = 0

    try:
        with conn.cursor() as cur:
            # Idempotência: remove dados anteriores deste manifest
            cur.execute(
                "DELETE FROM finep.projetos_operacao_direta WHERE manifest_id = %s",
                (manifest_id,),
            )
            deleted = cur.rowcount
            if deleted:
                logger.info(f"Removidos {deleted} registros anteriores do manifest_id={manifest_id}")

            # Carregar em batches
            batch = []
            for record in iter_rows(Path(xlsx_path)):
                record["manifest_id"] = manifest_id
                batch.append(record)

                if len(batch) >= batch_size:
                    psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                    rows_loaded += len(batch)
                    batch = []
                    logger.debug(f"  {rows_loaded} linhas inseridas...")

            # Último batch
            if batch:
                psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                rows_loaded += len(batch)

        conn.commit()
        logger.info(f"✓ {rows_loaded} linhas carregadas com sucesso")

    except Exception as exc:
        conn.rollback()
        logger.error(f"Falha na carga: {exc}")
        raise
    finally:
        conn.close()

    duration = (datetime.now(tz=timezone.utc) - t_start).total_seconds()

    # Atualizar manifesto
    mark_manifest_loaded(manifest_id, rows_loaded)

    return {
        "manifest_id":       manifest_id,
        "rows_loaded":       rows_loaded,
        "duration_seconds":  round(duration, 2),
        "xlsx_path":         xlsx_path,
    }
