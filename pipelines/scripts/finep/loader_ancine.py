"""
FINEP Loader — Projetos Ancine

Lê a aba Projetos_Ancine do arquivo Contratacao.xlsx mais recente
(ou do manifest_id informado) e popula finep.projetos_ancine.

Estrutura da aba:
  - ~175 linhas, 20 colunas
  - Header: linha 6 | Dados: linha 7..
  - Projetos audiovisuais do FSA (2009–2013)
  - Sem CNPJ Executor, sem Prazo Execução Original, sem Resumo Publicável

Estratégia de carga:
  - DELETE + INSERT por manifest_id: idempotente.
"""

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)

SHEET_NAME = "Projetos_Ancine"
HEADER_ROW = 6
DATA_START  = 7

INSERT_SQL = """
    INSERT INTO finep.projetos_ancine (
        instrumento, demanda, ref, contrato,
        data_assinatura, prazo_execucao,
        titulo, status,
        proponente, cnpj_proponente, executor,
        municipio, uf,
        valor_finep, contrapartida_financeira, contrapartida_nao_financeira,
        valor_pago, intervenientes,
        aporte_financeiro_interv, aporte_nao_financeiro_interv,
        manifest_id
    ) VALUES (
        %(instrumento)s, %(demanda)s, %(ref)s, %(contrato)s,
        %(data_assinatura)s, %(prazo_execucao)s,
        %(titulo)s, %(status)s,
        %(proponente)s, %(cnpj_proponente)s, %(executor)s,
        %(municipio)s, %(uf)s,
        %(valor_finep)s, %(contrapartida_financeira)s, %(contrapartida_nao_financeira)s,
        %(valor_pago)s, %(intervenientes)s,
        %(aporte_financeiro_interv)s, %(aporte_nao_financeiro_interv)s,
        %(manifest_id)s
    )
"""


# ---------------------------------------------------------------------------
# Banco de dados
# ---------------------------------------------------------------------------

def _get_db_connection():
    return psycopg2.connect(
        host=os.getenv("POSTGRES_HOST", "postgres"),
        port=int(os.getenv("POSTGRES_PORT", "5432")),
        database=os.getenv("POSTGRES_DB", "osint_metadata"),
        user=os.getenv("POSTGRES_USER", "osint_admin"),
        password=os.getenv("POSTGRES_PASSWORD", "osint_secure_password"),
    )


def get_latest_manifest(dataset_type: str = "contratacao") -> Optional[dict]:
    conn = _get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("""
                SELECT id, local_path, download_date
                FROM finep.download_manifest
                WHERE dataset_type = %s
                  AND processing_status IN ('downloaded', 'validated', 'loaded')
                  AND local_path IS NOT NULL
                ORDER BY download_date DESC
                LIMIT 1
            """, (dataset_type,))
            row = cur.fetchone()
            return dict(row) if row else None
    finally:
        conn.close()


def mark_manifest_loaded(manifest_id: int, rows_loaded: int) -> None:
    conn = _get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE finep.download_manifest
                SET processing_status = 'loaded',
                    rows_count        = COALESCE(rows_count, 0) + %s,
                    updated_at        = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (rows_loaded, manifest_id))
            conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_numeric(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _clean_int(val) -> Optional[int]:
    if val is None:
        return 0
    try:
        return int(val)
    except (TypeError, ValueError):
        return 0


def _clean_date(val):
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date()
    return None


# ---------------------------------------------------------------------------
# Leitura do xlsx
# ---------------------------------------------------------------------------

def iter_rows(xlsx_path: Path):
    """
    Lê a aba Projetos_Ancine e gera dicts prontos para INSERT.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            if all(v is None for v in row):
                continue

            record = {
                "instrumento":                  _clean_str(row[0]),
                "demanda":                      _clean_str(row[1]),
                "ref":                          _clean_str(row[2]),
                "contrato":                     _clean_str(row[3]),
                "data_assinatura":              _clean_date(row[4]),
                "prazo_execucao":               _clean_date(row[5]),
                "titulo":                       _clean_str(row[6]),
                "proponente":                   _clean_str(row[7]),
                "cnpj_proponente":              _clean_str(row[8]),
                "executor":                     _clean_str(row[9]),
                "municipio":                    _clean_str(row[10]),
                "uf":                           _clean_str(row[11]),
                "valor_finep":                  _clean_numeric(row[12]),
                "status":                       _clean_str(row[13]),
                "contrapartida_financeira":     _clean_numeric(row[14]),
                "contrapartida_nao_financeira": _clean_numeric(row[15]),
                "valor_pago":                   _clean_numeric(row[16]),
                "intervenientes":               _clean_int(row[17]),
                "aporte_financeiro_interv":     _clean_numeric(row[18]),
                "aporte_nao_financeiro_interv": _clean_numeric(row[19]) if len(row) > 19 else None,
                "manifest_id":                  None,
            }

            # Ignora linhas de separação sem instrumento
            if not record["instrumento"] and not record["contrato"]:
                continue

            yield record
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Carga principal
# ---------------------------------------------------------------------------

def load_ancine(
    manifest_id: Optional[int] = None,
    xlsx_path: Optional[str] = None,
    batch_size: int = 500,
) -> dict:
    """
    Carrega a aba Projetos_Ancine para o PostgreSQL.

    Args:
        manifest_id: ID do registro em finep.download_manifest.
        xlsx_path:   Caminho do arquivo xlsx.
        batch_size:  Linhas por batch de INSERT.

    Returns:
        dict com manifest_id, rows_loaded, duration_seconds
    """
    t_start = datetime.now(tz=timezone.utc)

    if manifest_id is None or xlsx_path is None:
        manifest = get_latest_manifest("contratacao")
        if not manifest:
            raise RuntimeError("Nenhum download de contratacao encontrado no manifesto.")
        manifest_id = manifest_id or manifest["id"]
        xlsx_path   = xlsx_path   or manifest["local_path"]

    logger.info(f"Carregando Projetos_Ancine | manifest_id={manifest_id} | {xlsx_path}")

    conn = _get_db_connection()
    rows_loaded = 0

    try:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM finep.projetos_ancine WHERE manifest_id = %s",
                (manifest_id,),
            )
            deleted = cur.rowcount
            if deleted:
                logger.info(f"Removidos {deleted} registros anteriores do manifest_id={manifest_id}")

            batch = []
            for record in iter_rows(Path(xlsx_path)):
                record["manifest_id"] = manifest_id
                batch.append(record)

                if len(batch) >= batch_size:
                    psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                    rows_loaded += len(batch)
                    batch = []
                    logger.debug(f"  {rows_loaded} linhas inseridas...")

            if batch:
                psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                rows_loaded += len(batch)

        conn.commit()
        logger.info(f"✓ {rows_loaded} linhas carregadas com sucesso")

    except Exception as exc:
        conn.rollback()
        logger.error(f"Falha na carga: {exc}")
        raise
    finally:
        conn.close()

    duration = (datetime.now(tz=timezone.utc) - t_start).total_seconds()
    mark_manifest_loaded(manifest_id, rows_loaded)

    return {
        "manifest_id":      manifest_id,
        "rows_loaded":      rows_loaded,
        "duration_seconds": round(duration, 2),
        "xlsx_path":        xlsx_path,
    }
