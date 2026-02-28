"""
FINEP Loader — Projetos Investimento

Lê a aba Projetos_Investimento do arquivo Contratacao.xlsx mais recente
(ou do manifest_id informado) e popula finep.projetos_investimento.

Estrutura da aba:
  - ~32 linhas, 19 colunas
  - Header: linha 6 | Dados: linha 7..
  - Instrumento de equity/venture capital com métricas de retorno (TIR, MOI)
  - Campos 'Sim'/'Não' são convertidos para BOOLEAN

Estratégia de carga:
  - DELETE + INSERT por manifest_id: idempotente.
"""

import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)

SHEET_NAME = "Projetos_Investimento"
HEADER_ROW = 6
DATA_START  = 7

INSERT_SQL = """
    INSERT INTO finep.projetos_investimento (
        ref, numero_contrato,
        cnpj_proponente, proponente,
        data_assinatura, data_follow_on,
        valor_follow_on, valor_total_contratado, valor_total_liberado,
        opcao_compra_exercida, opcao_compra_prorrogada,
        data_exercicio_opcao, valuation_opcao, participacao_finep,
        desinvestimento_realizado, data_desinvestimento, valor_desinvestimento,
        tir, moi,
        manifest_id
    ) VALUES (
        %(ref)s, %(numero_contrato)s,
        %(cnpj_proponente)s, %(proponente)s,
        %(data_assinatura)s, %(data_follow_on)s,
        %(valor_follow_on)s, %(valor_total_contratado)s, %(valor_total_liberado)s,
        %(opcao_compra_exercida)s, %(opcao_compra_prorrogada)s,
        %(data_exercicio_opcao)s, %(valuation_opcao)s, %(participacao_finep)s,
        %(desinvestimento_realizado)s, %(data_desinvestimento)s, %(valor_desinvestimento)s,
        %(tir)s, %(moi)s,
        %(manifest_id)s
    )
"""


# ---------------------------------------------------------------------------
# Banco de dados
# ---------------------------------------------------------------------------

def _get_db_connection():
    return psycopg2.connect(
        host=os.getenv("POSTGRES_HOST", "postgres"),
        port=int(os.getenv("POSTGRES_PORT", "5432")),
        database=os.getenv("POSTGRES_DB", "osint_metadata"),
        user=os.getenv("POSTGRES_USER", "osint_admin"),
        password=os.getenv("POSTGRES_PASSWORD", "osint_secure_password"),
    )


def get_latest_manifest(dataset_type: str = "contratacao") -> Optional[dict]:
    conn = _get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("""
                SELECT id, local_path, download_date
                FROM finep.download_manifest
                WHERE dataset_type = %s
                  AND processing_status IN ('downloaded', 'validated', 'loaded')
                  AND local_path IS NOT NULL
                ORDER BY download_date DESC
                LIMIT 1
            """, (dataset_type,))
            row = cur.fetchone()
            return dict(row) if row else None
    finally:
        conn.close()


def mark_manifest_loaded(manifest_id: int, rows_loaded: int) -> None:
    conn = _get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE finep.download_manifest
                SET processing_status = 'loaded',
                    rows_count        = COALESCE(rows_count, 0) + %s,
                    updated_at        = CURRENT_TIMESTAMP
                WHERE id = %s
            """, (rows_loaded, manifest_id))
            conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _clean_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_numeric(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _clean_date(val):
    if val is None:
        return None
    if hasattr(val, "date"):
        return val.date()
    return None


def _clean_bool(val) -> Optional[bool]:
    """Converte 'Sim' → True, 'Não'/'Nao' → False, None → None."""
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("sim", "yes", "1", "true"):
        return True
    if s in ("não", "nao", "no", "0", "false"):
        return False
    return None


# ---------------------------------------------------------------------------
# Leitura do xlsx
# ---------------------------------------------------------------------------

def iter_rows(xlsx_path: Path):
    """
    Lê a aba Projetos_Investimento e gera dicts prontos para INSERT.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            if all(v is None for v in row):
                continue

            record = {
                "ref":                         _clean_str(row[0]),
                "numero_contrato":             _clean_str(row[1]),
                "cnpj_proponente":             _clean_str(row[2]),
                "proponente":                  _clean_str(row[3]),
                "data_assinatura":             _clean_date(row[4]),
                "data_follow_on":              _clean_date(row[5]),
                "valor_follow_on":             _clean_numeric(row[6]),
                "valor_total_contratado":      _clean_numeric(row[7]),
                "valor_total_liberado":        _clean_numeric(row[8]),
                "opcao_compra_exercida":       _clean_bool(row[9]),
                "opcao_compra_prorrogada":     _clean_bool(row[10]),
                "data_exercicio_opcao":        _clean_date(row[11]),
                "valuation_opcao":             _clean_numeric(row[12]),
                "participacao_finep":          _clean_numeric(row[13]),
                "desinvestimento_realizado":   _clean_bool(row[14]),
                "data_desinvestimento":        _clean_date(row[15]),
                "valor_desinvestimento":       _clean_numeric(row[16]),
                "tir":                         _clean_numeric(row[17]),
                "moi":                         _clean_numeric(row[18]) if len(row) > 18 else None,
                "manifest_id":                 None,
            }

            # Ignora linhas sem número de contrato
            if not record["numero_contrato"]:
                continue

            yield record
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Carga principal
# ---------------------------------------------------------------------------

def load_investimento(
    manifest_id: Optional[int] = None,
    xlsx_path: Optional[str] = None,
    batch_size: int = 200,
) -> dict:
    """
    Carrega a aba Projetos_Investimento para o PostgreSQL.

    Args:
        manifest_id: ID do registro em finep.download_manifest.
        xlsx_path:   Caminho do arquivo xlsx.
        batch_size:  Linhas por batch (tabela pequena ~32 linhas).

    Returns:
        dict com manifest_id, rows_loaded, duration_seconds
    """
    t_start = datetime.now(tz=timezone.utc)

    if manifest_id is None or xlsx_path is None:
        manifest = get_latest_manifest("contratacao")
        if not manifest:
            raise RuntimeError("Nenhum download de contratacao encontrado no manifesto.")
        manifest_id = manifest_id or manifest["id"]
        xlsx_path   = xlsx_path   or manifest["local_path"]

    logger.info(f"Carregando Projetos_Investimento | manifest_id={manifest_id} | {xlsx_path}")

    conn = _get_db_connection()
    rows_loaded = 0

    try:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM finep.projetos_investimento WHERE manifest_id = %s",
                (manifest_id,),
            )
            deleted = cur.rowcount
            if deleted:
                logger.info(f"Removidos {deleted} registros anteriores do manifest_id={manifest_id}")

            batch = []
            for record in iter_rows(Path(xlsx_path)):
                record["manifest_id"] = manifest_id
                batch.append(record)

                if len(batch) >= batch_size:
                    psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                    rows_loaded += len(batch)
                    batch = []

            if batch:
                psycopg2.extras.execute_batch(cur, INSERT_SQL, batch, page_size=batch_size)
                rows_loaded += len(batch)

        conn.commit()
        logger.info(f"✓ {rows_loaded} linhas carregadas com sucesso")

    except Exception as exc:
        conn.rollback()
        logger.error(f"Falha na carga: {exc}")
        raise
    finally:
        conn.close()

    duration = (datetime.now(tz=timezone.utc) - t_start).total_seconds()
    mark_manifest_loaded(manifest_id, rows_loaded)

    return {
        "manifest_id":      manifest_id,
        "rows_loaded":      rows_loaded,
        "duration_seconds": round(duration, 2),
        "xlsx_path":        xlsx_path,
    }
