"""
FINEP Downloader

Baixa os arquivos Excel públicos disponibilizados pela FINEP
e registra cada coleta na tabela finep.download_manifest do PostgreSQL.

Arquivos rastreados:
  - http://download.finep.gov.br/Contratacao.xlsx
  - http://download.finep.gov.br/Liberacao.xlsx
"""

import hashlib
import logging
import os
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Optional

import psycopg2
import requests
from psycopg2.extras import DictCursor

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Datasets rastreados
# ---------------------------------------------------------------------------
FINEP_DATASETS = {
    "contratacao": "http://download.finep.gov.br/Contratacao.xlsx",
    "liberacao":   "http://download.finep.gov.br/Liberacao.xlsx",
}

# Timeout conservador, pois os arquivos podem ser grandes
REQUEST_TIMEOUT = (30, 300)   # (connect, read) em segundos


# ---------------------------------------------------------------------------
# Banco de dados
# ---------------------------------------------------------------------------

def _get_db_connection():
    """Abre conexão com o PostgreSQL via variáveis de ambiente."""
    return psycopg2.connect(
        host=os.getenv("POSTGRES_HOST", "postgres"),
        port=int(os.getenv("POSTGRES_PORT", "5432")),
        database=os.getenv("POSTGRES_DB", "osint_metadata"),
        user=os.getenv("POSTGRES_USER", "osint_admin"),
        password=os.getenv("POSTGRES_PASSWORD", "osint_secure_password"),
    )


# ---------------------------------------------------------------------------
# Utilitários HTTP
# ---------------------------------------------------------------------------

def _parse_last_modified(header_value: Optional[str]) -> Optional[datetime]:
    """Converte o header HTTP Last-Modified para datetime aware (UTC)."""
    if not header_value:
        return None
    try:
        return parsedate_to_datetime(header_value)
    except Exception:
        return None


def _sha256(path: Path) -> str:
    """Calcula SHA-256 de um arquivo de forma eficiente (buffer de 8 MB)."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8 * 1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Detecção de mudança via HEAD
# ---------------------------------------------------------------------------

def check_remote_headers(url: str) -> dict:
    """
    Faz uma requisição HEAD para detectar se o arquivo remoto mudou.

    Returns:
        dict com etag, last_modified, content_length
    """
    logger.info(f"HEAD request: {url}")
    resp = requests.head(url, timeout=REQUEST_TIMEOUT[0], allow_redirects=True)
    resp.raise_for_status()

    return {
        "etag":            resp.headers.get("ETag"),
        "last_modified":   _parse_last_modified(resp.headers.get("Last-Modified")),
        "content_length":  int(resp.headers["Content-Length"])
                           if "Content-Length" in resp.headers else None,
    }


def get_last_download_info(dataset_type: str) -> Optional[dict]:
    """
    Retorna as informações do último download bem-sucedido para o dataset.

    Returns:
        dict com etag, checksum, download_date ou None se não há histórico
    """
    conn = _get_db_connection()
    try:
        with conn.cursor(cursor_factory=DictCursor) as cur:
            cur.execute(
                """
                SELECT http_etag, file_checksum_sha256, download_date,
                       http_last_modified, http_content_length
                FROM finep.download_manifest
                WHERE dataset_type = %s
                  AND processing_status IN ('downloaded', 'validated')
                ORDER BY download_date DESC
                LIMIT 1
                """,
                (dataset_type,),
            )
            row = cur.fetchone()
            return dict(row) if row else None
    finally:
        conn.close()


def has_remote_changed(dataset_type: str, url: str) -> tuple[bool, dict]:
    """
    Verifica se o arquivo remoto mudou em relação ao último download.

    Returns:
        (changed: bool, remote_headers: dict)
    """
    remote = check_remote_headers(url)
    last = get_last_download_info(dataset_type)

    if last is None:
        logger.info(f"[{dataset_type}] Sem histórico — download necessário.")
        return True, remote

    # Comparar ETag (mais confiável quando disponível)
    if remote.get("etag") and last.get("http_etag"):
        changed = remote["etag"] != last["http_etag"]
        logger.info(
            f"[{dataset_type}] ETag comparado: "
            f"remoto={remote['etag']} vs local={last['http_etag']} → "
            f"{'MUDOU' if changed else 'igual'}"
        )
        return changed, remote

    # Fallback: Last-Modified
    if remote.get("last_modified") and last.get("http_last_modified"):
        changed = remote["last_modified"] > last["http_last_modified"]
        logger.info(
            f"[{dataset_type}] Last-Modified comparado: "
            f"remoto={remote['last_modified']} vs local={last['http_last_modified']} → "
            f"{'MUDOU' if changed else 'igual'}"
        )
        return changed, remote

    # Sem informação suficiente → baixar por precaução
    logger.warning(f"[{dataset_type}] Sem ETag nem Last-Modified — forçando download.")
    return True, remote


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def download_file(url: str, dest_path: Path) -> int:
    """
    Faz o download de uma URL para dest_path (streaming).

    Returns:
        Tamanho em bytes do arquivo salvo
    """
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info(f"Baixando {url} → {dest_path}")

    with requests.get(url, stream=True, timeout=REQUEST_TIMEOUT,
                      allow_redirects=True) as resp:
        resp.raise_for_status()
        total = 0
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8 * 1024 * 1024):
                f.write(chunk)
                total += len(chunk)

    logger.info(f"Download concluído: {total:,} bytes")
    return total


# ---------------------------------------------------------------------------
# Validação básica do Excel
# ---------------------------------------------------------------------------

def validate_xlsx(path: Path) -> dict:
    """
    Abre o arquivo Excel e extrai metadados básicos.

    Returns:
        dict com is_valid, rows_count, sheet_names, error_message
    """
    try:
        import openpyxl  # lazy import; não é necessário no script de download

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        # Conta linhas da primeira aba (excluindo cabeçalho)
        ws = wb.active
        rows_count = ws.max_row - 1 if ws.max_row and ws.max_row > 1 else 0
        wb.close()

        return {
            "is_valid":    True,
            "rows_count":  rows_count,
            "sheet_names": sheet_names,
            "error_message": None,
        }
    except Exception as exc:
        logger.error(f"Falha na validação do xlsx {path}: {exc}")
        return {
            "is_valid":      False,
            "rows_count":    None,
            "sheet_names":   None,
            "error_message": str(exc),
        }


# ---------------------------------------------------------------------------
# Registro no manifesto
# ---------------------------------------------------------------------------

def register_download(
    *,
    dataset_type: str,
    source_url: str,
    local_path: Path,
    file_size_bytes: int,
    checksum: str,
    remote_headers: dict,
    validation: dict,
) -> int:
    """
    Insere um registro na tabela finep.download_manifest.

    Returns:
        ID da linha inserida
    """
    status = "validated" if validation["is_valid"] else "failed"
    file_name = local_path.name

    conn = _get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO finep.download_manifest (
                    file_name, source_url, dataset_type,
                    download_date,
                    file_size_bytes, file_checksum_sha256, local_path,
                    http_last_modified, http_etag, http_content_length,
                    is_valid, rows_count, sheet_names,
                    processing_status, error_message,
                    updated_at
                ) VALUES (
                    %s, %s, %s,
                    %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    CURRENT_TIMESTAMP
                )
                RETURNING id
                """,
                (
                    file_name,
                    source_url,
                    dataset_type,
                    datetime.now(tz=timezone.utc),
                    file_size_bytes,
                    checksum,
                    str(local_path),
                    remote_headers.get("last_modified"),
                    remote_headers.get("etag"),
                    remote_headers.get("content_length"),
                    validation["is_valid"],
                    validation["rows_count"],
                    validation["sheet_names"],
                    status,
                    validation["error_message"],
                ),
            )
            row_id = cur.fetchone()[0]
            conn.commit()
            logger.info(f"[{dataset_type}] Manifesto registrado — id={row_id}, status={status}")
            return row_id
    except Exception as exc:
        conn.rollback()
        logger.error(f"Falha ao registrar no manifesto: {exc}")
        raise
    finally:
        conn.close()


def register_failed(
    *,
    dataset_type: str,
    source_url: str,
    error_message: str,
) -> None:
    """Registra uma tentativa de download que falhou."""
    file_name = source_url.rstrip("/").split("/")[-1]
    conn = _get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO finep.download_manifest (
                    file_name, source_url, dataset_type,
                    download_date, processing_status, error_message,
                    updated_at
                ) VALUES (%s, %s, %s, %s, 'failed', %s, CURRENT_TIMESTAMP)
                """,
                (
                    file_name,
                    source_url,
                    dataset_type,
                    datetime.now(tz=timezone.utc),
                    error_message,
                ),
            )
            conn.commit()
            logger.warning(f"[{dataset_type}] Falha registrada no manifesto.")
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Função principal de coleta de um dataset
# ---------------------------------------------------------------------------

def collect_dataset(
    dataset_type: str,
    url: str,
    output_dir: Path,
    force: bool = False,
) -> dict:
    """
    Orquestra a coleta completa de um dataset FINEP:
    1. Verifica se o arquivo remoto mudou (HEAD)
    2. Faz o download se necessário
    3. Calcula checksum
    4. Valida o xlsx
    5. Registra no manifesto

    Args:
        dataset_type: 'contratacao' ou 'liberacao'
        url:          URL de download
        output_dir:   Diretório local de destino
        force:        Se True, baixa mesmo que não haja mudança detectada

    Returns:
        dict com resultado da coleta
    """
    logger.info(f"=== Coletando dataset: {dataset_type} ===")

    # 1. Verificar se mudou
    try:
        changed, remote_headers = has_remote_changed(dataset_type, url)
    except Exception as exc:
        logger.warning(f"[{dataset_type}] Falha no HEAD request: {exc} — forçando download.")
        changed = True
        remote_headers = {}

    if not changed and not force:
        logger.info(f"[{dataset_type}] Arquivo inalterado — download ignorado.")
        return {"dataset_type": dataset_type, "skipped": True, "reason": "unchanged"}

    # 2. Download
    file_name = url.rstrip("/").split("/")[-1]
    timestamp = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M%S")
    dest_path = output_dir / dataset_type / f"{timestamp}_{file_name}"

    try:
        file_size = download_file(url, dest_path)
    except Exception as exc:
        error_msg = f"Download falhou: {exc}"
        logger.error(f"[{dataset_type}] {error_msg}")
        register_failed(dataset_type=dataset_type, source_url=url, error_message=error_msg)
        raise

    # 3. Checksum
    checksum = _sha256(dest_path)
    logger.info(f"[{dataset_type}] SHA-256: {checksum}")

    # 4. Validação
    validation = validate_xlsx(dest_path)
    if not validation["is_valid"]:
        logger.warning(f"[{dataset_type}] Arquivo inválido: {validation['error_message']}")

    # 5. Registro
    manifest_id = register_download(
        dataset_type=dataset_type,
        source_url=url,
        local_path=dest_path,
        file_size_bytes=file_size,
        checksum=checksum,
        remote_headers=remote_headers,
        validation=validation,
    )

    return {
        "dataset_type":  dataset_type,
        "skipped":       False,
        "manifest_id":   manifest_id,
        "local_path":    str(dest_path),
        "file_size_bytes": file_size,
        "checksum":      checksum,
        "rows_count":    validation["rows_count"],
        "sheet_names":   validation["sheet_names"],
        "is_valid":      validation["is_valid"],
    }


def collect_all(output_dir: Path, force: bool = False) -> list[dict]:
    """
    Coleta todos os datasets FINEP definidos em FINEP_DATASETS.

    Returns:
        Lista de resultados por dataset
    """
    results = []
    for dataset_type, url in FINEP_DATASETS.items():
        try:
            result = collect_dataset(dataset_type, url, output_dir, force=force)
            results.append(result)
        except Exception as exc:
            logger.error(f"Falha na coleta de {dataset_type}: {exc}")
            results.append({
                "dataset_type": dataset_type,
                "skipped":      False,
                "error":        str(exc),
            })
    return results
